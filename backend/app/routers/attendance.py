from datetime import datetime, timedelta, date, time
from io import StringIO
import csv
from typing import Dict, List, Tuple
from uuid import UUID
import httpx
from collections import defaultdict

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.config import get_settings
from app.database import get_db
from app.models import (
    AttendanceBoardThreshold,
    AttendanceDashboardExport,
    AttendanceEvent,
    AttendanceFaceTemplate,
    AttendanceSessionConfig,
    Building,
    ClassSession,
    Enrollment,
    Floor,
    Room,
    Subject,
    Student,
    User,
)
from app.routers.auth import get_current_user, get_user_room_scope
from app.schemas.common import (
    AttendanceConfigUpsert,
    AttendanceDailyRoomSummary,
    AttendanceEventIngest,
    AttendanceDashboardBreakdownPoint,
    AttendanceDashboardBreakdownResponse,
    AttendanceDashboardKpiResponse,
    AttendanceDashboardRankingResponse,
    AttendanceDashboardRankingRow,
    AttendanceDashboardTrendPoint,
    AttendanceDashboardTrendResponse,
    AttendanceMockEventIngest,
    AttendanceSessionReport,
    AttendanceStudentHistoryEntry,
    AttendanceStudentStatus,
)

router = APIRouter(prefix="/api/attendance", tags=["Attendance"])
settings = get_settings()


def _attendance_stream_target(path: str) -> str:
    base_url = settings.attendance_service_url.rstrip("/")
    normalized_path = path if path.startswith("/") else f"/{path}"
    return f"{base_url}{normalized_path}"


async def _proxy_attendance_json(path: str) -> Response:
    target_url = _attendance_stream_target(path)
    timeout = httpx.Timeout(connect=2.0, read=2.0, write=2.0, pool=2.0)

    try:
        async with httpx.AsyncClient(timeout=timeout) as client:
            upstream = await client.get(target_url)
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=503, detail=f"Attendance stream service unavailable ({exc.__class__.__name__})") from exc

    content_type = upstream.headers.get("content-type", "application/json")
    return Response(content=upstream.content, status_code=upstream.status_code, headers={"Content-Type": content_type})


@router.get("/stream/status")
async def get_attendance_stream_status(current_user: User = Depends(get_current_user)):
    """Proxy attendance stream status for frontend clients via backend origin."""
    _ = current_user
    return await _proxy_attendance_json("/status")


@router.get("/stream/health")
async def get_attendance_stream_health(current_user: User = Depends(get_current_user)):
    """Proxy attendance stream health for frontend clients via backend origin."""
    _ = current_user
    return await _proxy_attendance_json("/health")


@router.get("/stream/video_feed")
async def get_attendance_stream_video_feed(current_user: User = Depends(get_current_user)):
    """Proxy MJPEG camera stream through backend so frontend stays on same origin."""
    _ = current_user
    target_url = _attendance_stream_target("/video_feed")
    timeout = httpx.Timeout(connect=3.0, read=30.0, write=10.0, pool=10.0)
    client = httpx.AsyncClient(timeout=timeout)

    try:
        request = client.build_request("GET", target_url)
        upstream = await client.send(request, stream=True)
    except httpx.HTTPError as exc:
        await client.aclose()
        raise HTTPException(status_code=503, detail=f"Attendance stream service unavailable ({exc.__class__.__name__})") from exc

    if upstream.status_code >= 400:
        payload = await upstream.aread()
        content_type = upstream.headers.get("content-type", "text/plain")
        status_code = upstream.status_code
        await upstream.aclose()
        await client.aclose()
        return Response(content=payload, status_code=status_code, headers={"Content-Type": content_type})

    content_type = upstream.headers.get("content-type", "multipart/x-mixed-replace; boundary=frame")

    async def iter_stream():
        try:
            async for chunk in upstream.aiter_raw():
                if chunk:
                    yield chunk
        finally:
            await upstream.aclose()
            await client.aclose()

    return StreamingResponse(
        iter_stream(),
        headers={
            "Content-Type": content_type,
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


def _ensure_attendance_role(current_user: User) -> None:
    if current_user.role not in {"LECTURER", "SYSTEM_ADMIN"}:
        raise HTTPException(status_code=403, detail="Only LECTURER or SYSTEM_ADMIN can access attendance APIs")


def _ensure_attendance_dashboard_role(current_user: User) -> None:
    if current_user.role not in {"SYSTEM_ADMIN", "ACADEMIC_BOARD", "FACILITY_STAFF", "LECTURER"}:
        raise HTTPException(status_code=403, detail="Insufficient role for school-wide attendance dashboard")


def _ensure_attendance_scope(current_user: User, room_id: UUID, db: Session) -> None:
    if current_user.role == "SYSTEM_ADMIN":
        return

    allowed_rooms = set(get_user_room_scope(current_user, db))
    if room_id not in allowed_rooms:
        raise HTTPException(status_code=403, detail="User not assigned to this room")


def _get_session_or_404(db: Session, session_id: UUID) -> ClassSession:
    session = db.query(ClassSession).filter(ClassSession.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    return session


def _get_or_create_attendance_config(db: Session, session_id: UUID) -> AttendanceSessionConfig:
    config = db.query(AttendanceSessionConfig).filter(AttendanceSessionConfig.session_id == session_id).first()
    if config:
        return config

    config = AttendanceSessionConfig(session_id=session_id)
    db.add(config)
    db.commit()
    db.refresh(config)
    return config


def _get_enrolled_students(db: Session, subject_id: UUID) -> List[Student]:
    return (
        db.query(Student)
        .join(Enrollment, Enrollment.student_id == Student.id)
        .filter(Enrollment.subject_id == subject_id)
        .all()
    )


def _get_first_recognized_event_map(
    db: Session,
    session_id: UUID,
    min_confidence: float,
) -> Dict[UUID, AttendanceEvent]:
    events = (
        db.query(AttendanceEvent)
        .filter(
            AttendanceEvent.session_id == session_id,
            AttendanceEvent.is_recognized.is_(True),
            AttendanceEvent.face_confidence >= min_confidence,
        )
        .order_by(AttendanceEvent.occurred_at.asc())
        .all()
    )

    first_seen: Dict[UUID, AttendanceEvent] = {}
    for event in events:
        if event.student_id not in first_seen:
            first_seen[event.student_id] = event
    return first_seen


def _derive_student_statuses(
    session: ClassSession,
    config: AttendanceSessionConfig,
    enrolled_students: List[Student],
    first_seen_map: Dict[UUID, AttendanceEvent],
) -> Tuple[List[AttendanceStudentStatus], Dict[str, int]]:
    cutoff = session.start_time + timedelta(minutes=config.grace_minutes)
    items: List[AttendanceStudentStatus] = []
    totals = {"present": 0, "late": 0, "absent": 0, "enrolled": len(enrolled_students)}

    for student in enrolled_students:
        first_event = first_seen_map.get(student.id)

        if first_event is None:
            status = "ABSENT"
            totals["absent"] += 1
        elif first_event.occurred_at <= cutoff:
            status = "PRESENT"
            totals["present"] += 1
        else:
            status = "LATE"
            totals["late"] += 1

        items.append(
            AttendanceStudentStatus(
                student_id=student.id,
                student_code=student.student_id,
                student_name=student.name,
                status=status,
                first_seen_at=first_event.occurred_at if first_event else None,
                confidence=first_event.face_confidence if first_event else None,
            )
        )

    return items, totals


def _safe_attendance_rate(enrolled: int, present: int, late: int) -> float:
    if enrolled <= 0:
        return 0.0
    return round(((present + late) / enrolled) * 100.0, 2)


def _weekday_label(value: int) -> str:
    mapping = {
        0: "Mon",
        1: "Tue",
        2: "Wed",
        3: "Thu",
        4: "Fri",
        5: "Sat",
        6: "Sun",
    }
    return mapping.get(value, str(value))


def _iso_week_key(input_date: date) -> str:
    year, week, _ = input_date.isocalendar()
    return f"{year}-W{week:02d}"


def _month_key(input_date: date) -> str:
    return f"{input_date.year}-{input_date.month:02d}"


def _build_session_attendance_rows(
    db: Session,
    sessions: List[ClassSession],
) -> tuple[List[dict], Dict[str, str], Dict[str, str], Dict[str, str], Dict[str, date]]:
    if not sessions:
        return [], {}, {}, {}, {}

    session_ids = [session.id for session in sessions]
    subject_ids = {session.subject_id for session in sessions if session.subject_id is not None}

    configs = (
        db.query(AttendanceSessionConfig)
        .filter(AttendanceSessionConfig.session_id.in_(session_ids))
        .all()
    )
    config_by_session = {config.session_id: config for config in configs}

    enrollments = (
        db.query(Enrollment.subject_id, Enrollment.student_id)
        .filter(Enrollment.subject_id.in_(subject_ids))
        .all()
        if subject_ids
        else []
    )
    students_by_subject: Dict[UUID, set[UUID]] = defaultdict(set)
    for subject_id, student_id in enrollments:
        students_by_subject[subject_id].add(student_id)

    events = (
        db.query(AttendanceEvent)
        .filter(
            AttendanceEvent.session_id.in_(session_ids),
            AttendanceEvent.is_recognized.is_(True),
        )
        .order_by(AttendanceEvent.occurred_at.asc())
        .all()
    )

    earliest_seen: Dict[tuple[UUID, UUID], datetime] = {}
    for event in events:
        session_config = config_by_session.get(event.session_id)
        threshold = session_config.min_confidence if session_config else 0.75
        if event.face_confidence < threshold:
            continue

        key = (event.session_id, event.student_id)
        if key not in earliest_seen:
            earliest_seen[key] = event.occurred_at

    room_labels: Dict[str, str] = {}
    subject_labels: Dict[str, str] = {}
    session_labels: Dict[str, str] = {}
    session_dates: Dict[str, date] = {}
    rows: List[dict] = []

    for session in sessions:
        if session.subject_id is None:
            continue

        roster = students_by_subject.get(session.subject_id, set())
        enrolled = len(roster)
        session_config = config_by_session.get(session.id)
        grace_minutes = session_config.grace_minutes if session_config else 10
        cutoff = session.start_time + timedelta(minutes=grace_minutes)

        present = 0
        late = 0
        absent = 0

        for student_id in roster:
            first_seen = earliest_seen.get((session.id, student_id))
            if first_seen is None:
                absent += 1
            elif first_seen <= cutoff:
                present += 1
            else:
                late += 1

        session_key = str(session.id)
        room_key = str(session.room_id)
        subject_key = str(session.subject_id)
        weekday = session.start_time.weekday()

        room_label = session.room.room_code if session.room and session.room.room_code else room_key[:8]
        subject_label = session.subject.code if session.subject and session.subject.code else (session.subject.name if session.subject else subject_key[:8])
        session_label = f"{room_label} | {session.start_time.strftime('%H:%M')} | {session_key[:8]}"

        room_labels[room_key] = room_label
        subject_labels[subject_key] = subject_label
        session_labels[session_key] = session_label
        session_dates[session_key] = session.start_time.date()

        rows.append(
            {
                "session_id": session_key,
                "session_label": session_label,
                "room_id": room_key,
                "room_label": room_label,
                "subject_id": subject_key,
                "subject_label": subject_label,
                "weekday": weekday,
                "weekday_label": _weekday_label(weekday),
                "start_time": session.start_time,
                "session_status": session.status,
                "enrolled": enrolled,
                "present": present,
                "late": late,
                "absent": absent,
                "attendance_rate": _safe_attendance_rate(enrolled, present, late),
            }
        )

    return rows, room_labels, subject_labels, session_labels, session_dates


def _apply_attendance_scope_filters(
    query,
    current_user: User,
    db: Session,
    building_id: UUID | None,
    room_id: UUID | None,
):
    allowed_rooms = get_user_room_scope(current_user, db)
    if allowed_rooms:
        query = query.filter(ClassSession.room_id.in_(allowed_rooms))

    if room_id:
        query = query.filter(ClassSession.room_id == room_id)

    if building_id:
        query = query.join(Room, Room.id == ClassSession.room_id).join(Floor, Floor.id == Room.floor_id).filter(Floor.building_id == building_id)

    return query


def _load_dashboard_sessions(
    db: Session,
    current_user: User,
    start_date: date | None,
    end_date: date | None,
    building_id: UUID | None,
    room_id: UUID | None,
    subject_id: UUID | None,
    session_id: UUID | None,
    day_of_week: int | None,
) -> List[ClassSession]:
    query = db.query(ClassSession).join(Room, Room.id == ClassSession.room_id).join(Floor, Floor.id == Room.floor_id)

    query = _apply_attendance_scope_filters(query, current_user, db, building_id, room_id)

    if subject_id:
        query = query.filter(ClassSession.subject_id == subject_id)
    if session_id:
        query = query.filter(ClassSession.id == session_id)

    if start_date:
        query = query.filter(ClassSession.start_time >= datetime.combine(start_date, time.min))
    if end_date:
        query = query.filter(ClassSession.start_time <= datetime.combine(end_date, time.max))

    sessions = query.order_by(ClassSession.start_time.asc()).all()

    if day_of_week is not None:
        sessions = [session for session in sessions if session.start_time.weekday() == day_of_week]

    return sessions


def _aggregate_rows(rows: List[dict]) -> Dict[str, int | float]:
    enrolled = sum(row["enrolled"] for row in rows)
    present = sum(row["present"] for row in rows)
    late = sum(row["late"] for row in rows)
    absent = sum(row["absent"] for row in rows)
    return {
        "enrolled": enrolled,
        "present": present,
        "late": late,
        "absent": absent,
        "attendance_rate": _safe_attendance_rate(enrolled, present, late),
    }


def _group_rows_for_breakdown(rows: List[dict], dimension: str) -> List[AttendanceDashboardBreakdownPoint]:
    grouped: Dict[str, Dict[str, int]] = defaultdict(lambda: {"enrolled": 0, "present": 0, "late": 0, "absent": 0})
    labels: Dict[str, str] = {}

    for row in rows:
        if dimension == "session":
            group_key = row["session_id"]
            label = row["session_label"]
        elif dimension == "subject":
            group_key = row["subject_id"]
            label = row["subject_label"]
        else:
            group_key = str(row["weekday"])
            label = row["weekday_label"]

        grouped[group_key]["enrolled"] += row["enrolled"]
        grouped[group_key]["present"] += row["present"]
        grouped[group_key]["late"] += row["late"]
        grouped[group_key]["absent"] += row["absent"]
        labels[group_key] = label

    items: List[AttendanceDashboardBreakdownPoint] = []
    for key, totals in grouped.items():
        items.append(
            AttendanceDashboardBreakdownPoint(
                key=key,
                label=labels.get(key, key),
                enrolled=totals["enrolled"],
                present=totals["present"],
                late=totals["late"],
                absent=totals["absent"],
                attendance_rate=_safe_attendance_rate(totals["enrolled"], totals["present"], totals["late"]),
            )
        )

    if dimension == "day_of_week":
        items.sort(key=lambda item: int(item.key))
    else:
        items.sort(key=lambda item: item.label)

    return items


def _group_rows_for_trend(rows: List[dict], granularity: str) -> List[AttendanceDashboardTrendPoint]:
    grouped: Dict[str, Dict[str, int]] = defaultdict(lambda: {"enrolled": 0, "present": 0, "late": 0, "absent": 0})
    labels: Dict[str, str] = {}

    for row in rows:
        start_time = row["start_time"]
        if granularity == "week":
            key = _iso_week_key(start_time.date())
            label = key
        elif granularity == "month":
            key = _month_key(start_time.date())
            label = start_time.strftime("%b %Y")
        elif granularity == "weekday":
            key = str(row["weekday"])
            label = row["weekday_label"]
        else:
            key = start_time.date().isoformat()
            label = start_time.strftime("%d %b")

        grouped[key]["enrolled"] += row["enrolled"]
        grouped[key]["present"] += row["present"]
        grouped[key]["late"] += row["late"]
        grouped[key]["absent"] += row["absent"]
        labels[key] = label

    points: List[AttendanceDashboardTrendPoint] = []
    for key, totals in grouped.items():
        points.append(
            AttendanceDashboardTrendPoint(
                key=key,
                label=labels.get(key, key),
                enrolled=totals["enrolled"],
                present=totals["present"],
                late=totals["late"],
                absent=totals["absent"],
                attendance_rate=_safe_attendance_rate(totals["enrolled"], totals["present"], totals["late"]),
            )
        )

    if granularity == "weekday":
        points.sort(key=lambda item: int(item.key))
    else:
        points.sort(key=lambda item: item.key)

    return points


def _build_dashboard_filters_csv(
    start_date: date | None,
    end_date: date | None,
    building_id: UUID | None,
    room_id: UUID | None,
    subject_id: UUID | None,
    session_id: UUID | None,
    day_of_week: int | None,
) -> str:
    return (
        f"start_date={start_date or ''},end_date={end_date or ''},building_id={building_id or ''},"
        f"room_id={room_id or ''},subject_id={subject_id or ''},session_id={session_id or ''},day_of_week={day_of_week if day_of_week is not None else ''}"
    )


def _build_dashboard_filters_payload(
    start_date: date | None,
    end_date: date | None,
    building_id: UUID | None,
    room_id: UUID | None,
    subject_id: UUID | None,
    session_id: UUID | None,
    day_of_week: int | None,
) -> dict:
    return {
        "start_date": start_date.isoformat() if start_date else None,
        "end_date": end_date.isoformat() if end_date else None,
        "building_id": str(building_id) if building_id else None,
        "room_id": str(room_id) if room_id else None,
        "subject_id": str(subject_id) if subject_id else None,
        "session_id": str(session_id) if session_id else None,
        "day_of_week": day_of_week,
    }


def _resolve_threshold_target(db: Session, building_id: UUID | None) -> float:
    if building_id:
        building = db.query(Building).filter(Building.id == building_id).first()
        if building and building.code:
            scoped = (
                db.query(AttendanceBoardThreshold)
                .filter(
                    AttendanceBoardThreshold.scope_type == "BUILDING",
                    AttendanceBoardThreshold.scope_id == building.code,
                )
                .first()
            )
            if scoped:
                return round(scoped.min_attendance_rate, 2)

    school_default = (
        db.query(AttendanceBoardThreshold)
        .filter(
            AttendanceBoardThreshold.scope_type == "SCHOOL",
            AttendanceBoardThreshold.scope_id == "GLOBAL",
        )
        .first()
    )
    if school_default:
        return round(school_default.min_attendance_rate, 2)

    return 85.0


def _record_dashboard_export(
    db: Session,
    current_user: User,
    export_format: str,
    row_count: int,
    filters_payload: dict,
    status: str,
    failure_reason: str | None = None,
) -> None:
    try:
        db.add(
            AttendanceDashboardExport(
                requested_by=current_user.id,
                export_format=export_format,
                filter_payload=filters_payload,
                row_count=row_count,
                status=status,
                failure_reason=failure_reason,
            )
        )
        db.commit()
    except Exception:
        db.rollback()


@router.get("/dashboard/kpis", response_model=AttendanceDashboardKpiResponse)
async def get_attendance_dashboard_kpis(
    start_date: date | None = Query(default=None),
    end_date: date | None = Query(default=None),
    building_id: UUID | None = Query(default=None),
    room_id: UUID | None = Query(default=None),
    subject_id: UUID | None = Query(default=None),
    session_id: UUID | None = Query(default=None),
    day_of_week: int | None = Query(default=None, ge=0, le=6),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_attendance_dashboard_role(current_user)

    sessions = _load_dashboard_sessions(
        db,
        current_user,
        start_date,
        end_date,
        building_id,
        room_id,
        subject_id,
        session_id,
        day_of_week,
    )
    rows, *_ = _build_session_attendance_rows(db, sessions)
    aggregate = _aggregate_rows(rows)
    aggregate["target_attendance_rate"] = _resolve_threshold_target(db, building_id)

    return AttendanceDashboardKpiResponse(**aggregate)


@router.get("/dashboard/breakdown", response_model=AttendanceDashboardBreakdownResponse)
async def get_attendance_dashboard_breakdown(
    dimension: str = Query(default="day_of_week", pattern="^(session|day_of_week|subject)$"),
    start_date: date | None = Query(default=None),
    end_date: date | None = Query(default=None),
    building_id: UUID | None = Query(default=None),
    room_id: UUID | None = Query(default=None),
    subject_id: UUID | None = Query(default=None),
    session_id: UUID | None = Query(default=None),
    day_of_week: int | None = Query(default=None, ge=0, le=6),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_attendance_dashboard_role(current_user)

    sessions = _load_dashboard_sessions(
        db,
        current_user,
        start_date,
        end_date,
        building_id,
        room_id,
        subject_id,
        session_id,
        day_of_week,
    )
    rows, *_ = _build_session_attendance_rows(db, sessions)
    points = _group_rows_for_breakdown(rows, dimension)

    return AttendanceDashboardBreakdownResponse(dimension=dimension, points=points)


@router.get("/dashboard/trend", response_model=AttendanceDashboardTrendResponse)
async def get_attendance_dashboard_trend(
    granularity: str = Query(default="weekday", pattern="^(day|week|month|weekday)$"),
    start_date: date | None = Query(default=None),
    end_date: date | None = Query(default=None),
    building_id: UUID | None = Query(default=None),
    room_id: UUID | None = Query(default=None),
    subject_id: UUID | None = Query(default=None),
    session_id: UUID | None = Query(default=None),
    day_of_week: int | None = Query(default=None, ge=0, le=6),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_attendance_dashboard_role(current_user)

    sessions = _load_dashboard_sessions(
        db,
        current_user,
        start_date,
        end_date,
        building_id,
        room_id,
        subject_id,
        session_id,
        day_of_week,
    )
    rows, *_ = _build_session_attendance_rows(db, sessions)
    points = _group_rows_for_trend(rows, granularity)

    return AttendanceDashboardTrendResponse(granularity=granularity, points=points)


@router.get("/dashboard/rankings", response_model=AttendanceDashboardRankingResponse)
async def get_attendance_dashboard_rankings(
    scope: str = Query(default="session", pattern="^(session|room|subject)$"),
    sort: str = Query(default="desc", pattern="^(asc|desc)$"),
    top_n: int = Query(default=20, ge=1, le=100),
    start_date: date | None = Query(default=None),
    end_date: date | None = Query(default=None),
    building_id: UUID | None = Query(default=None),
    room_id: UUID | None = Query(default=None),
    subject_id: UUID | None = Query(default=None),
    session_id: UUID | None = Query(default=None),
    day_of_week: int | None = Query(default=None, ge=0, le=6),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_attendance_dashboard_role(current_user)

    sessions = _load_dashboard_sessions(
        db,
        current_user,
        start_date,
        end_date,
        building_id,
        room_id,
        subject_id,
        session_id,
        day_of_week,
    )
    rows, room_labels, subject_labels, _, _ = _build_session_attendance_rows(db, sessions)

    if scope == "session":
        ranking_base = []
        for row in rows:
            ranking_base.append(
                {
                    "scope_key": row["session_id"],
                    "scope_label": row["session_label"],
                    "start_time": row["start_time"],
                    "session_status": row["session_status"],
                    "enrolled": row["enrolled"],
                    "present": row["present"],
                    "late": row["late"],
                    "absent": row["absent"],
                    "attendance_rate": row["attendance_rate"],
                }
            )
    else:
        grouped: Dict[str, Dict[str, int]] = defaultdict(lambda: {"enrolled": 0, "present": 0, "late": 0, "absent": 0})
        for row in rows:
            key = row["room_id"] if scope == "room" else row["subject_id"]
            grouped[key]["enrolled"] += row["enrolled"]
            grouped[key]["present"] += row["present"]
            grouped[key]["late"] += row["late"]
            grouped[key]["absent"] += row["absent"]

        ranking_base = []
        for key, totals in grouped.items():
            label = room_labels.get(key, key) if scope == "room" else subject_labels.get(key, key)
            ranking_base.append(
                {
                    "scope_key": key,
                    "scope_label": label,
                    "start_time": None,
                    "session_status": None,
                    "enrolled": totals["enrolled"],
                    "present": totals["present"],
                    "late": totals["late"],
                    "absent": totals["absent"],
                    "attendance_rate": _safe_attendance_rate(totals["enrolled"], totals["present"], totals["late"]),
                }
            )

    reverse = sort == "desc"
    ranking_base.sort(key=lambda row: row["attendance_rate"], reverse=reverse)

    rows_out: List[AttendanceDashboardRankingRow] = []
    for index, row in enumerate(ranking_base[:top_n], start=1):
        rows_out.append(AttendanceDashboardRankingRow(rank=index, **row))

    return AttendanceDashboardRankingResponse(scope=scope, rows=rows_out)


@router.get("/dashboard/export")
async def export_attendance_dashboard(
    format: str = Query(default="xlsx", pattern="^(xlsx|csv)$"),
    start_date: date | None = Query(default=None),
    end_date: date | None = Query(default=None),
    building_id: UUID | None = Query(default=None),
    room_id: UUID | None = Query(default=None),
    subject_id: UUID | None = Query(default=None),
    session_id: UUID | None = Query(default=None),
    day_of_week: int | None = Query(default=None, ge=0, le=6),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_attendance_dashboard_role(current_user)

    sessions = _load_dashboard_sessions(
        db,
        current_user,
        start_date,
        end_date,
        building_id,
        room_id,
        subject_id,
        session_id,
        day_of_week,
    )
    rows, _, _, _, _ = _build_session_attendance_rows(db, sessions)
    filters_payload = _build_dashboard_filters_payload(start_date, end_date, building_id, room_id, subject_id, session_id, day_of_week)

    kpis = _aggregate_rows(rows)
    breakdown_points = _group_rows_for_breakdown(rows, "day_of_week")
    trend_points = _group_rows_for_trend(rows, "weekday")

    ranking_rows = sorted(rows, key=lambda row: row["attendance_rate"], reverse=True)
    filter_text = _build_dashboard_filters_csv(start_date, end_date, building_id, room_id, subject_id, session_id, day_of_week)

    if format == "csv":
        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["ATTENDANCE DASHBOARD EXPORT"])
        writer.writerow(["filters", filter_text])
        writer.writerow(["generated_at", datetime.utcnow().isoformat()])
        writer.writerow([])
        writer.writerow(["KPI", "value"])
        writer.writerow(["enrolled", kpis["enrolled"]])
        writer.writerow(["present", kpis["present"]])
        writer.writerow(["late", kpis["late"]])
        writer.writerow(["absent", kpis["absent"]])
        writer.writerow(["attendance_rate", kpis["attendance_rate"]])
        writer.writerow([])
        writer.writerow(["Breakdown by weekday"])
        writer.writerow(["label", "enrolled", "present", "late", "absent", "attendance_rate"])
        for point in breakdown_points:
            writer.writerow([point.label, point.enrolled, point.present, point.late, point.absent, point.attendance_rate])
        writer.writerow([])
        writer.writerow(["Ranking by session"])
        writer.writerow(["session", "start_time", "status", "enrolled", "present", "late", "absent", "attendance_rate"])
        for row in ranking_rows:
            writer.writerow([
                row["session_label"],
                row["start_time"].isoformat() if row["start_time"] else "",
                row["session_status"],
                row["enrolled"],
                row["present"],
                row["late"],
                row["absent"],
                row["attendance_rate"],
            ])

        _record_dashboard_export(
            db,
            current_user,
            export_format="CSV",
            row_count=len(ranking_rows),
            filters_payload=filters_payload,
            status="SUCCESS",
        )

        return Response(
            content=buffer.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=attendance-dashboard-export.csv"},
        )

    try:
        from openpyxl import Workbook  # type: ignore[import-not-found]
    except Exception as exc:
        _record_dashboard_export(
            db,
            current_user,
            export_format="XLSX",
            row_count=0,
            filters_payload=filters_payload,
            status="FAILED",
            failure_reason="openpyxl dependency unavailable",
        )
        raise HTTPException(
            status_code=503,
            detail="XLSX export dependency is unavailable. Install openpyxl in backend environment.",
        ) from exc

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["Enrolled", kpis["enrolled"]])
    summary_sheet.append(["Present", kpis["present"]])
    summary_sheet.append(["Late", kpis["late"]])
    summary_sheet.append(["Absent", kpis["absent"]])
    summary_sheet.append(["Attendance Rate", kpis["attendance_rate"]])
    summary_sheet.append(["Generated At", datetime.utcnow().isoformat()])
    summary_sheet.append(["Filters", filter_text])

    breakdown_sheet = workbook.create_sheet("Breakdown")
    breakdown_sheet.append(["Label", "Enrolled", "Present", "Late", "Absent", "Attendance Rate"])
    for point in breakdown_points:
        breakdown_sheet.append([point.label, point.enrolled, point.present, point.late, point.absent, point.attendance_rate])

    trend_sheet = workbook.create_sheet("Trend")
    trend_sheet.append(["Label", "Enrolled", "Present", "Late", "Absent", "Attendance Rate"])
    for point in trend_points:
        trend_sheet.append([point.label, point.enrolled, point.present, point.late, point.absent, point.attendance_rate])

    ranking_sheet = workbook.create_sheet("Ranking")
    ranking_sheet.append(["Session", "Start Time", "Status", "Enrolled", "Present", "Late", "Absent", "Attendance Rate"])
    for row in ranking_rows:
        ranking_sheet.append([
            row["session_label"],
            row["start_time"].isoformat() if row["start_time"] else "",
            row["session_status"],
            row["enrolled"],
            row["present"],
            row["late"],
            row["absent"],
            row["attendance_rate"],
        ])

    _record_dashboard_export(
        db,
        current_user,
        export_format="XLSX",
        row_count=len(ranking_rows),
        filters_payload=filters_payload,
        status="SUCCESS",
    )

    from io import BytesIO
    bytes_io = BytesIO()
    workbook.save(bytes_io)

    return Response(
        content=bytes_io.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=attendance-dashboard-export.xlsx"},
    )


@router.put("/sessions/{session_id}/config", response_model=AttendanceConfigUpsert)
async def upsert_session_attendance_config(
    session_id: UUID,
    payload: AttendanceConfigUpsert,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_attendance_role(current_user)

    session = _get_session_or_404(db, session_id)
    _ensure_attendance_scope(current_user, session.room_id, db)

    config = _get_or_create_attendance_config(db, session_id)
    config.grace_minutes = payload.grace_minutes
    config.min_confidence = payload.min_confidence
    config.auto_checkin_enabled = payload.auto_checkin_enabled

    db.commit()

    return AttendanceConfigUpsert(
        grace_minutes=config.grace_minutes,
        min_confidence=config.min_confidence,
        auto_checkin_enabled=config.auto_checkin_enabled,
    )


@router.post("/sessions/{session_id}/events/mock")
async def ingest_mock_attendance_event(
    session_id: UUID,
    payload: AttendanceMockEventIngest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_attendance_role(current_user)

    session = _get_session_or_404(db, session_id)
    _ensure_attendance_scope(current_user, session.room_id, db)

    if session.subject_id is None:
        raise HTTPException(status_code=400, detail="Session subject is required for attendance")

    enrollment_exists = (
        db.query(Enrollment)
        .filter(Enrollment.subject_id == session.subject_id, Enrollment.student_id == payload.student_id)
        .first()
    )
    if not enrollment_exists:
        raise HTTPException(status_code=400, detail="Student is not enrolled in this session subject")

    config = _get_or_create_attendance_config(db, session_id)
    is_recognized = config.auto_checkin_enabled and payload.face_confidence >= config.min_confidence

    event = AttendanceEvent(
        session_id=session_id,
        student_id=payload.student_id,
        source=payload.source,
        face_confidence=payload.face_confidence,
        is_recognized=is_recognized,
        occurred_at=payload.occurred_at or datetime.utcnow(),
        event_metadata=payload.metadata,
        created_by_user_id=current_user.id,
    )

    db.add(event)
    db.commit()
    db.refresh(event)

    return {
        "event_id": event.id,
        "session_id": session_id,
        "student_id": payload.student_id,
        "recognized": event.is_recognized,
        "grace_minutes": config.grace_minutes,
        "min_confidence": config.min_confidence,
    }


@router.post("/sessions/{session_id}/events/ingest")
async def ingest_attendance_event(
    session_id: UUID,
    payload: AttendanceEventIngest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Real attendance event ingest endpoint.
    Called by the PC attendance service (USB webcam face recognition).
    Accepts LECTURER, SYSTEM_ADMIN, or EXAM_PROCTOR roles.
    """
    if current_user.role not in {"LECTURER", "SYSTEM_ADMIN", "EXAM_PROCTOR"}:
        raise HTTPException(status_code=403, detail="Insufficient role for attendance ingest")

    session = _get_session_or_404(db, session_id)

    if session.subject_id is None:
        raise HTTPException(status_code=400, detail="Session subject is required for attendance")

    # Resolve student by student_id
    student = db.query(Student).filter(Student.id == payload.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    # Verify enrollment
    enrollment_exists = (
        db.query(Enrollment)
        .filter(Enrollment.subject_id == session.subject_id, Enrollment.student_id == payload.student_id)
        .first()
    )
    if not enrollment_exists:
        raise HTTPException(status_code=400, detail="Student is not enrolled in this session subject")

    config = _get_or_create_attendance_config(db, session_id)
    is_recognized = config.auto_checkin_enabled and payload.face_confidence >= config.min_confidence

    event = AttendanceEvent(
        session_id=session_id,
        student_id=payload.student_id,
        source=payload.source or "USB_WEBCAM",
        face_confidence=payload.face_confidence,
        is_recognized=is_recognized,
        occurred_at=payload.occurred_at or datetime.utcnow(),
        event_metadata=payload.metadata or {},
        created_by_user_id=current_user.id,
    )

    db.add(event)
    db.commit()
    db.refresh(event)

    # Derive current status for this student
    cutoff = session.start_time + timedelta(minutes=config.grace_minutes)
    if event.occurred_at <= cutoff:
        status = "PRESENT"
    else:
        status = "LATE"

    return {
        "event_id": str(event.id),
        "session_id": str(session_id),
        "student_id": str(payload.student_id),
        "student_code": student.student_id,
        "student_name": student.name,
        "recognized": event.is_recognized,
        "status": status,
        "grace_minutes": config.grace_minutes,
        "min_confidence": config.min_confidence,
    }


@router.get("/face-templates/students")
async def list_face_template_students(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    List all students that have active face templates.
    Used by the attendance service to build student_code -> student_id mapping.
    """
    _ensure_attendance_role(current_user)

    students_with_templates = (
        db.query(Student)
        .join(AttendanceFaceTemplate, AttendanceFaceTemplate.student_id == Student.id)
        .filter(AttendanceFaceTemplate.is_active.is_(True))
        .distinct()
        .all()
    )

    # Also include all students even without templates (for the PC service mapping)
    all_students = db.query(Student).all()

    return [
        {
            "student_id": str(s.id),
            "student_code": s.student_id,
            "name": s.name,
            "has_template": any(t.student_id == s.id for t in students_with_templates) if students_with_templates else False,
        }
        for s in all_students
    ]


@router.get("/sessions/{session_id}", response_model=AttendanceSessionReport)
async def get_session_attendance_report(
    session_id: UUID,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_attendance_role(current_user)

    session = _get_session_or_404(db, session_id)
    _ensure_attendance_scope(current_user, session.room_id, db)

    if session.subject_id is None:
        raise HTTPException(status_code=400, detail="Session subject is required for attendance")

    config = _get_or_create_attendance_config(db, session_id)
    enrolled_students = _get_enrolled_students(db, session.subject_id)
    first_seen_map = _get_first_recognized_event_map(db, session_id, config.min_confidence)

    statuses, totals = _derive_student_statuses(session, config, enrolled_students, first_seen_map)

    return AttendanceSessionReport(
        session_id=session.id,
        room_id=session.room_id,
        room_code=session.room.room_code if session.room else None,
        start_time=session.start_time,
        end_time=session.end_time,
        grace_minutes=config.grace_minutes,
        min_confidence=config.min_confidence,
        totals=totals,
        students=statuses,
    )


@router.get("/sessions/{session_id}/export")
async def export_session_attendance_csv(
    session_id: UUID,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    report = await get_session_attendance_report(session_id=session_id, current_user=current_user, db=db)

    csv_buffer = StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow(["student_id", "student_code", "student_name", "status", "first_seen_at", "confidence"])
    for item in report.students:
        writer.writerow(
            [
                str(item.student_id),
                item.student_code,
                item.student_name,
                item.status,
                item.first_seen_at.isoformat() if item.first_seen_at else "",
                item.confidence if item.confidence is not None else "",
            ]
        )

    filename = f"attendance-session-{session_id}.csv"
    return Response(
        content=csv_buffer.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/students/{student_id}/history", response_model=List[AttendanceStudentHistoryEntry])
async def get_student_attendance_history(
    student_id: UUID,
    limit: int = Query(default=30, ge=1, le=365),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_attendance_role(current_user)

    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    sessions = (
        db.query(ClassSession)
        .join(Enrollment, Enrollment.subject_id == ClassSession.subject_id)
        .filter(Enrollment.student_id == student_id)
        .order_by(ClassSession.start_time.desc())
        .limit(limit)
        .all()
    )

    results: List[AttendanceStudentHistoryEntry] = []
    for session in sessions:
        _ensure_attendance_scope(current_user, session.room_id, db)
        config = _get_or_create_attendance_config(db, session.id)
        first_seen = (
            db.query(AttendanceEvent)
            .filter(
                AttendanceEvent.session_id == session.id,
                AttendanceEvent.student_id == student_id,
                AttendanceEvent.is_recognized.is_(True),
                AttendanceEvent.face_confidence >= config.min_confidence,
            )
            .order_by(AttendanceEvent.occurred_at.asc())
            .first()
        )

        cutoff = session.start_time + timedelta(minutes=config.grace_minutes)
        if first_seen is None:
            status = "ABSENT"
        elif first_seen.occurred_at <= cutoff:
            status = "PRESENT"
        else:
            status = "LATE"

        results.append(
            AttendanceStudentHistoryEntry(
                session_id=session.id,
                subject_id=session.subject_id,
                room_id=session.room_id,
                start_time=session.start_time,
                end_time=session.end_time,
                status=status,
                first_seen_at=first_seen.occurred_at if first_seen else None,
            )
        )

    return results


@router.get("/rooms/{room_id}/daily-summary", response_model=AttendanceDailyRoomSummary)
async def get_room_daily_attendance_summary(
    room_id: UUID,
    day: date | None = Query(default=None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_attendance_role(current_user)
    _ensure_attendance_scope(current_user, room_id, db)

    day = day or date.today()

    start_dt = datetime.combine(day, time.min)
    end_dt = datetime.combine(day, time.max)

    sessions = (
        db.query(ClassSession)
        .filter(
            ClassSession.room_id == room_id,
            ClassSession.start_time >= start_dt,
            ClassSession.start_time <= end_dt,
        )
        .all()
    )

    totals = {"present": 0, "late": 0, "absent": 0, "enrolled": 0}

    for session in sessions:
        if session.subject_id is None:
            continue
        config = _get_or_create_attendance_config(db, session.id)
        enrolled_students = _get_enrolled_students(db, session.subject_id)
        first_seen_map = _get_first_recognized_event_map(db, session.id, config.min_confidence)
        _, session_totals = _derive_student_statuses(session, config, enrolled_students, first_seen_map)

        totals["present"] += session_totals["present"]
        totals["late"] += session_totals["late"]
        totals["absent"] += session_totals["absent"]
        totals["enrolled"] += session_totals["enrolled"]

    return AttendanceDailyRoomSummary(
        room_id=room_id,
        date=day,
        sessions_count=len(sessions),
        totals=totals,
    )
